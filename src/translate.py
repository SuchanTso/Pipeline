import pandas as pd
import yaml
import openpyxl
from collections import defaultdict
import os
import numpy as np
from datetime import datetime, timedelta
import json
import re
import argparse


def parse_args():
    """
    解析命令行参数
    """
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "-i",
        "--input_file",
        type=str,
        default="data/whole_Haishu.xlsx",
        help="path of the input data file",
    )
    parser.add_argument(
        "-o",
        "--output_file",
        type=str,
        default="data/whole_Haishu_output.xlsx",
        help="sheet of output data file",
    )
    parser.add_argument(
        "--config",
        type=str,
        default="config/translate.yaml",
        help="path of the configuration file",
    )
    
    return parser.parse_args()
def is_date_format(num_format):
    """检查数字格式是否为日期格式"""
    if num_format is None:
        return False
    
    # Excel日期格式常用代码
    date_codes = ['yy', 'yyyy', 'mm', 'dd', 'hh', 'ss', 'am/pm', 'a/p', 
                  '年', '月', '日', '时', '分', '秒', 'e', 'g']
    
    # 检查格式是否包含日期相关代码
    return any(code in num_format.lower() for code in date_codes)

def is_date_cell(cell):
    """通过单元格格式判断是否为日期单元格"""
    return isinstance(cell.value,(pd.Timestamp, datetime))

def convert_excel_date(serial_date, num_format):
    """将Excel日期序列号转换为日期字符串"""
    try:
        base_date = datetime(1900, 1, 1)
        if serial_date > 60:
            serial_date -= 1
            
        days = int(serial_date)
        fraction = serial_date - days
        
        dt = base_date + timedelta(days=days-1)
        
        seconds = int(fraction * 86400)
        hours, remainder = divmod(seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        
        if hours > 0 or minutes > 0 or seconds > 0:
            dt = dt.replace(hour=hours, minute=minutes, second=seconds)
        
        if num_format is None:
            return dt.strftime('%Y-%m-%d')
        
        if 'h' in num_format.lower() or 'hh' in num_format.lower():
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        else:
            return dt.strftime('%Y-%m-%d')
            
    except Exception as e:
        print(f"日期转换错误: {serial_date} -> {e}")
        return str(serial_date)

def get_cell_value(cell, is_header=False):
    """获取单元格的值，智能处理日期和坐标值"""
    value = cell.value
    
    # 处理空值
    if value is None:
        return ""
    
    # 处理日期单元格
    if not is_header and is_date_cell(cell):
        if isinstance(value, (int, float)):
            # 确保是有效日期序列号
            if 0 < value < 3000000:  # 1900到9999年范围
                return convert_excel_date(value, cell.number_format)
        elif isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')
    
    # 处理其他数值
    if isinstance(value, float) and not is_header:
        # 坐标值通常有小数部分，但整数也不会影响
        return value
    
    # 处理布尔值
    if isinstance(value, bool):
        return str(value).lower()
    
    # 默认返回字符串
    return str(value) if value is not None else ""

def load_content_mappings(mapping_files):
    """从JSON文件加载内容映射规则"""
    content_mappings = {}
    for mapping_file in mapping_files:
        column = mapping_file['column']
        path = mapping_file['path']
        
        if not os.path.exists(path):
            print(f"警告: 内容映射文件不存在: {path}")
            continue
        
        try:
            with open(path, 'r', encoding='utf-8') as f:
                mapping_data = json.load(f)
                content_mappings[column] = mapping_data
                # print(f"已加载 {column} 的内容映射: {len(mapping_data)} 条规则")
        except Exception as e:
            print(f"加载内容映射文件 {path} 时出错: {e}")
    
    return content_mappings

def process_excel_with_config(input_excel, output_excel, config_file, sheet_name=None):
    """处理Excel文件，使用配置文件进行表头映射和内容替换"""
    # 读取YAML配置文件
    with open(config_file, 'r', encoding='utf-8') as f:
        config = yaml.safe_load(f)
    
    # 获取映射规则
    header_mapping = config.get('header_mapping', {})
    content_mapping_files = config.get('content_mapping_files', [])
    
    # 加载内容映射规则
    content_mappings = load_content_mappings(content_mapping_files)
    
    # 读取整个Excel文件结构
    original_wb = openpyxl.load_workbook(input_excel, data_only=False)
    
    # 确定要处理的工作表
    if sheet_name:
        sheet_names = [sheet_name]
    else:
        sheet_names = original_wb.sheetnames
    
    # 创建新的工作簿
    new_wb = openpyxl.Workbook()
    new_wb.remove(new_wb.active)
    
    for sheet_name in sheet_names:
        # print(f"\n处理工作表: {sheet_name}")
        
        # 读取原始数据
        if sheet_name in original_wb.sheetnames:
            original_sheet = original_wb[sheet_name]
        else:
            print(f"警告: 工作表 '{sheet_name}' 不存在，跳过")
            continue
        
        # 读取表头行
        headers = []
        header_row = list(original_sheet.iter_rows(min_row=1, max_row=1))[0]
        for cell in header_row:
            headers.append(get_cell_value(cell, is_header=True))
        
        # 读取数据行
        data_rows = []
        date_columns = set()  # 存储被识别为日期的列
        
        # 标记日期列
        for col_idx in range(1, len(headers) + 1):
            cell = original_sheet.cell(row=2, column=col_idx)  # 检查第一行数据
            if is_date_cell(cell):
                date_columns.add(headers[col_idx - 1])
                # print(f"检测到日期列: {headers[col_idx - 1]}")
        
        for row in original_sheet.iter_rows(min_row=2):
            row_data = []
            for cell in row:
                # 使用智能值获取
                row_data.append(get_cell_value(cell))
            data_rows.append(row_data)
        
        df = pd.DataFrame(data_rows, columns=headers)
        
        # 记录原始列名重复情况
        header_counts = defaultdict(int)
        unique_columns = []
        col_position_map = {}
        
        # 处理重复列名
        for i, col in enumerate(df.columns):
            header_counts[col] += 1
            count = header_counts[col]
            
            if count > 1:
                unique_col = f"{i}_{col}"
            else:
                unique_col = col
                
            unique_columns.append(unique_col)
            col_position_map[unique_col] = (i, col)
        
        # 应用唯一标识作为临时列名
        df.columns = unique_columns
        
        # 应用表头映射
        new_columns = []
        col_mapping = {}
        reverse_col_mapping = {}
        
        for unique_col in unique_columns:
            orig_index, orig_col = col_position_map[unique_col]
            position_key = f"{orig_index}_{orig_col}"
            
            if position_key in header_mapping:
                new_header = header_mapping[position_key]
            elif orig_col in header_mapping:
                new_header = header_mapping[orig_col]
            else:
                new_header = orig_col
            
            if new_header in new_columns:
                suffix = 1
                while f"{new_header}_{suffix}" in new_columns:
                    suffix += 1
                new_header = f"{new_header}_{suffix}"
            
            new_columns.append(new_header)
            col_mapping[unique_col] = new_header
            reverse_col_mapping[new_header] = unique_col
        
        # 应用新列名
        df.columns = new_columns
        
        # 应用内容映射
        # print("\n应用内容映射:")
        processed_columns = set()
        
        for new_col, mapping_dict in content_mappings.items():
            # 检查列是否存在
            if new_col not in df.columns:
                # 尝试使用原始列名映射
                unique_col = reverse_col_mapping.get(new_col)
                if not unique_col or unique_col not in col_position_map:
                    # print(f"  - 列 '{new_col}' 不存在于工作表中，跳过")
                    continue
                
                orig_index, orig_col = col_position_map[unique_col]
                # print(f"  - 映射列 '{new_col}' ({orig_col})")
            else:
                pass
                # print(f"  - 映射列 '{new_col}'")
            
            # 应用映射规则
            df[new_col] = df[new_col].apply(
                lambda x: mapping_dict.get(str(x), mapping_dict.get(x, x)))
            processed_columns.add(new_col)
        
        # 将处理后的数据写入新工作表
        new_sheet = new_wb.create_sheet(title=sheet_name)
        new_sheet.append(list(df.columns))
        
        # 写入数据行
        for _, row in df.iterrows():
            row_vals = list(row)
            new_sheet.append(row_vals)
        
        # 设置新工作表的格式（保持坐标值为数值格式）
        for col_idx, col_name in enumerate(df.columns, 1):
            # 对于坐标列，保持数值格式
            if "X(" in col_name or "Y(" in col_name or "坐标" in col_name:
                for row_idx in range(2, len(df) + 2):
                    cell = new_sheet.cell(row=row_idx, column=col_idx)
                    if isinstance(cell.value, float) or isinstance(cell.value, int):
                        cell.number_format = "0.000"
                        cell.value = float(cell.value)
            
            # 对于日期列，设置日期格式
            elif col_name in date_columns:
                for row_idx in range(2, len(df) + 2):
                    cell = new_sheet.cell(row=row_idx, column=col_idx)
                    if isinstance(cell.value, str) and re.match(r"\d{4}-\d{2}-\d{2}", cell.value):
                        cell.number_format = "yyyy-mm-dd"
        
        # 创建处理报告
        report = {
            "sheet_name": sheet_name,
            "total_rows": len(df),
            "total_columns": len(df.columns),
            "renamed_headers": sum(1 for old, new in col_mapping.items() if old != new),
            "content_mapped_columns": list(processed_columns),
            "date_columns_identified": list(date_columns),
            "coordinate_columns": [
                col for col in new_columns 
                if "X(" in col or "Y(" in col or "坐标" in col
            ]
        }
        
        # 保存报告
        report_path = os.path.splitext(output_excel)[0] + f"_{sheet_name}_report.yaml"
        with open(report_path, 'w', encoding='utf-8') as f:
            yaml.dump(report, f, allow_unicode=True, sort_keys=False)
        
        # print(f"工作表 {sheet_name} 处理完成，报告保存至: {report_path}")
    
    # 保存整个工作簿
    new_wb.save(output_excel)
    print(f"\nExcel文件处理完成，保存至: {output_excel}")

if __name__ == "__main__":
    args = parse_args()
    # 示例使用
    process_excel_with_config(
        input_excel=args.input_file,
        output_excel=args.output_file,
        config_file=args.config
    )